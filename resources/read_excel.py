import openpyxl

def get_row_count(file,sheet_name):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    return (sheet.max_row)


def get_column_count(file,sheet_name):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    return(sheet.max_column)


def get_data(file,sheet_name,row_num,col_num):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheet_name)
    return sheet.cell(row=row_num,column=col_num).value

